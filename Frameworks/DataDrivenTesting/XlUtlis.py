import openpyxl

def getrowcount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)

def getcolumncount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)

def readdatafile(path,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=colnum).value

def writedatafile(path,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=colnum).value = data
    workbook.save(path)

