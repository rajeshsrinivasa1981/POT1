import openpyxl

path="C:\Selenium\Excel_Write1.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("Sheet1")

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(2,rows+1):
    sheet.cell(row=r,column=4,value="India")

workbook.save(path)