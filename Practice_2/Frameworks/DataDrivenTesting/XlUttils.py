import openpyxl

#Get Row count
def getrowcount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return(sheet.max_row)

#Get Column Count
def getcolumncount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return(sheet.max_column)

#Read data into excel

def readdata(path,sheetname,rownum,columnnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnnum).value

#write data into excel

def writedata(path,sheetname,rownum, columnnum,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    sheet.cell(row=rownum,column=columnnum,value=data)
    workbook.save(path)






