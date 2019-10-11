"""
#READ EXCEL FILE 
import openpyxl

path="C:\Learning\Selenium\DDT\Excel_Write_Read.xlsx" #Excel path
workbook=openpyxl.load_workbook(path) #Create Workbook

sheet=workbook["Sheet1"] #Create Sheet

rows=sheet.max_row #rows count
columns=sheet.max_column #columns count

print(rows)
print(columns)

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(row=r,column=c).value,end='   ')
    print()
"""

""" 
#Write Data into Excel File
import openpyxl
path="C:\Learning\Selenium\DDT\Excel_Write_Read.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet2"]

for r in range(1,4):
    for c in range(1,3):
        sheet.cell(row=r,column=c,value="Lokesh")
        #sheet.cell(row=r,column=c).value="Rajesh"

workbook.save(path)
"""









